import logging
import os
import re
import tkinter as tk
from datetime import datetime, time, timedelta
from tkinter import filedialog, messagebox, simpledialog

from openpyxl import load_workbook, Workbook
from openpyxl.utils.datetime import from_excel
from rapidfuzz import process, fuzz
# Configuração do logging
logging.basicConfig(level=logging.INFO)

# Definição das listas pré-definidas no nível global
clientes_predefinidos = [
    "Adara Franceline",
    "Ademar Mendes Bezerra Junior",
    "Adna Ribeiro",
    "Adriana Lobão",
    "Adriana Martins Sá",
    "Adriana Teofila",
    "Adriano Abintes",
    "Agapito Feitosa",
    "Aglair Aguiar",
    "Águeda Muniz",
    "Albina Luciana Gonçalves",
    "Alcides Ferreira Rego Neto",
    "Alcides Saldanha Lima",
    "Alex Viana",
    "Alexandre Agostune",
    "Alexandre Henrique Silva",
    "Alexandre Medeiros",
    "Alexandre Teixeira",
    "Alexandre Varjal",
    "Aline Pires de Freitas",
    "Amanda Cavalcante Martins Soares",
    "Amanda Costa da Silva",
    "Amanda Gomes Albuquerque",
    "Amanda Thaynara Ferreira",
    "Amanda Zélia de Sousa Tavares",
    "Ana Beatriz Montenegro",
    "Ana Claudia Vidal",
    "Ana Dara Carolina",
    "Ana Karla Gomes Pereira",
    "Ana Karolina Sales",
    "Ana Laura Costa Menezes",
    "Ana Luiza Ximenes Dias",
    "Ana Marilia Maia Magalhães",
    "Ana Marques",
    "Ana Taisa Barbosa de Mendonça",
    "Ana Vitoria Delfino",
    "Anderson Costa",
    "Andre Costa",
    "André Luiz Meireles Justi",
    "Andre Oliveira Sena",
    "Andre Rios",
    "Andre Sousa Castelo",
    "Andrei Manuel Coutinho Leite",
    "Andreia Silveira",
    "Andressa Crisitna Gonçalves",
    "Anna Sofia Ribeira Mesquita",
    "Antonio Ferreira Da Costa Neto",
    "Antonio Roberto Vila Nova",
    "Antonio Victor Sales de Carvalho",
    "Augusto Lima",
    "Augusto Vieira",
    "Barbara Fernandes Malafaia",
    "Beatriz Furtado",
    "Beatriz Leite Jereissate",
    "Brena Maria Almeida",
    "Bruna Paula da Silva Santos",
    "Bruna Weyne Germano",
    "Camila Almeida",
    "Camila Falcão Catunda",
    "Camila Ladislau",
    "Camila Ventura",
    "Camila Verissimo Medeiros",
    "Camille Macedo Albuquerque",
    "Carla Castelo Branco",
    "Carla Eufrasio",
    "Carlos Enrique Correia Xavier",
    "Carolina Araripe",
    "Carolina Barreira Diogenes",
    "Cassia Carol",
    "Cauê Espanhol",
    "Cecilia Barroso De Oliveira",
    "Cesar Juca",
    "Cintia Alves",
    "Cirnia Cabral Alves",
    "Clarissa Quevedo",
    "Clarisse Uchoa de Albuquerque",
    "Claryanne Aguiar",
    "Claudia Brilhante",
    "Claudia Pompeu",
    "Claudia Talita",
    "Claudia Veras",
    "Claudio Albuquerque Sales",
    "Cláudio Regis Sampaio",
    "Clemilcelia Karan",
    "Cristiane Alexandre",
    "Cyntia Ferreira Gomes",
    "Dalmiro Castro",
    "Dani Brito",
    "Daniela Pinheiro Gomes",
    "Daniela Rolim Medeiros",
    "Danielle de Paula Magalhães",
    "Danielly Coutinho",
    "Danilo Barbosa Pimentel",
    "Danilo Gurgel Mota",
    "Dany Alcantara",
    "Dario Palácio Mesquita",
    "Davi Dantas",
    "David Montenegro",
    "Diana Montenegro",
    "Diego Cabral Ferreira",
    "Diego Nogueira Kaur",
    "Diogo Souza Brito Alencar",
    "Dione Cardoso",
    "Edilvania Duarte",
    "Eduarda Rabelo",
    "Eduardo Almeida",
    "Eduardo Barreira",
    "Eduardo Brasileiro",
    "Eduardo Vitorio",
    "Eladio Bede Filho",
    "Eliana Gomes",
    "Elisangela Cavalcante Costa",
    "Emanuel Sátiro",
    "Emanuele Scolare",
    "Emerson de Primio",
    "Emilse Ximenes",
    "Eri Johnson Vieira Laurentino",
    "Ezequias Da Silva Leite",
    "Fabiana Diogenes",
    "Fabiana Espanhol",
    "Fabiola Feijó",
    "Felipe Araujo",
    "Felipe Lavor",
    "Felipe Pena Forte",
    "Felipe Souza Pinheiro",
    "Felipe TESTE",
    "Fernanda Gomes",
    "Fernanda Pinho",
    "Fernanda Vasconcelos",
    "Fernando Frota Sampaio",
    "Fernando Negreiros",
    "Francelarne De Paula",
    "Francisca Manuela Carvalho Gomes",
    "Francisco Eufrasio",
    "Francisco Ian De Vasconcelos",
    "Francisco Jereissati Neto",
    "Fred Castro",
    "Frederico Ferreira",
    "Gabriela Balduino Brasil",
    "Gabriela Parente",
    "Gabriele Gutierrez Rocha de Oliveira",
    "Gaspar Oliveira",
    "Geordany Rosé de Oliveira Viana Esmeraldo",
    "Georgelia Carvalho",
    "Gisele Reis",
    "Giselly Costa Lima",
    "Glaucia Barbosa",
    "Glayciane Lima",
    "Gonçalo Souto Diogo Junior",
    "Graziela Costa Lima Maia",
    "Guilherme Vilanova",
    "Gustavo Dorea Carneiro",
    "Henrique Marinho",
    "Hugo Pinheiro",
    "Iara Silva Dias",
    "Iarley Italo Alves da Silva",
    "Igor Jarlles",
    "Inês Gurgel",
    "Ingrid Cardozo Botelho",
    "Ingrid Louback de Moura",
    "Ingrid Paula Nogueira da Silva",
    "Isabela Lomonaco",
    "Isabele Siqueira Melo",
    "Isabella Vasconcelos",
    "Isabelle Teixeira Loureiro",
    "Israel Dantas",
    "Italo Bastos",
    "Ivana Dantas",
    "Ivana Maria Gomes Cavalcante",
    "Izabel Campelo",
    "Jackson Gomes Costa Da Fonseca",
    "Jade Catunda",
    "James TESTE HUB85",
    "Jamile Maia Braide",
    "Jamilly Brito",
    "JB TESTE",
    "Jelma Guimarães",
    "Jessica Alves",
    "Jessica Sampaio",
    "Jhonny Makis Gomes",
    "Joane Aguinezio",
    "João Manoel Braollos",
    "Jorge Luiz Passamani",
    "Joriza Magalhães",
    "José Anderson Araujo",
    "José Teles",
    "José Victor Lima Figuereido",
    "Joyce Alves Marques",
    "Juliana Rezende Dias Coelho",
    "Juliana Solheiro",
    "Juracir Mourão",
    "Karine Mesquita",
    "Karisa Carolina Teixeira",
    "Karlla Jordão",
    "Kelly Arruda",
    "Kelly Mesias",
    "Laise Linhares",
    "Laisse Rodrigues Linhares",
    "Larissa Luciane Campelo Diógenes",
    "Lásaro Henrique Lopes Gonçalves",
    "Laura Bounin",
    "Leandro Moreira",
    "Leda Castelar",
    "Leila Silveira Vieira",
    "Leticia Morais",
    "Liliana Maria Mota Moreira",
    "Livia do Vale",
    "Lourdes Nogueira",
    "Lourene de Arruda",
    "Luana Lucia Ferreira De Oliveira",
    "Luana Uchoa",
    "Lucas Bruno Borges",
    "Lucas Farias Freire",
    "Lucas Grangeiro de Castro",
    "Lucas Marinho",
    "Lucas Vieira",
    "Luciana Farias",
    "Luciana Leite",
    "Luciana Lima Pinto",
    "Luciane Maas Da Silva",
    "Luciane Vieto Marques",
    "Lucimar Castro Sousa",
    "Luiz Cesar Cardoso Gomes",
    "Luiz Edson Correia Sales",
    "Lurdes Negreiros",
    "Maiara Rocha",
    "Maina Araujo Sampaio",
    "Manoela Viana De Albuquerque",
    "Manuel Decio Pinheiro Neto",
    "Manuela Bezerril Fernandes",
    "Manuela Sales Barroso",
    "Manuela Silveira",
    "Mara Gabryelle Dias Ribeiro Rodrigues",
    "Marcela Castelar",
    "Marcelo Cavalcante",
    "Marcelo Pereira Gonçalves",
    "Marcelo Siqueira",
    "Marcio Fraga",
    "Margarida Bana",
    "Maria Angelucia",
    "Maria Carolina Pontes",
    "Maria Clara Soares Mapurunga",
    "Maria de Fátima de Oliveira",
    "Maria de Lourdes Silva Lima de Sousa",
    "Maria do Socorro Calixto",
    "Maria Efigenia Bezerra",
    "Maria Ferreira",
    "Maria Helena Sad",
    "Maria Leticia Araujo",
    "Maria Luiza Pinheiro",
    "Mariana Gonçalves Ferreira",
    "Mariana Vasconcelos",
    "Marilia De Sousa Carneiro",
    "Marília Gomes Cruz",
    "Marina Leal Luna Silva",
    "Mario Queiroz",
    "Mark Kevin",
    "Marlena Chaves",
    "Maryana Ferreira Cardoso",
    "Mateus Feitosa",
    "Matheus Facó Jesuino Simões",
    "Matheus Valença",
    "Mayara Cordeiro",
    "Mayara Dos Santos Rodrigues",
    "Melinda de Paula",
    "Michela Militana Fernandes Pinheiro",
    "MIlena Daniel Fontenele de Sousa",
    "Mirella Genezini Marotte",
    "Monique Kitayama",
    "Naiana Cunha",
    "Natalia Sena", 'Natacha campos arriada' ,
    "Nathalin da Silva Pinto Câmara",
    "Nelson Vidal",
    "Ogresio Mores",
    "Oscar da Fonte",
    "Otilia Pessoa",
    "Pablo Amorim",
    "Pablo Luan Lopes",
    "Patricia Sousa De Morais Tarja",
    "Patricia Telles",
    "Paula Machado Couto",
    "Paula Ramalho",
    "Paulo Alexandre Leite Gomes",
    "Paulo Cezar de Oliveira",
    "Pedrette TESTE",
    "Pedro Ciqueira",
    "Pedro Elpidio Gadelha Guimarães Pinheiro",
    "Pedro Gerson de Amorim",
    "Persivo Ribeiro",
    "Poliana Rachid",
    "Priscila Cavalcanti",
    "Priscilla Falcão",
    "Rafael Coimbra Fernandes",
    "Rafael Lima Freitas Guimarães",
    "Rafael Souza",
    "Raidel Teixeira",
    "Raissa Ivo",
    "Raissa Julia",
    "Raniere Franco",
    "Raphael Fabricio De Andrade Sales",
    "Raquel Meneses de Moura",
    "Raquel Queiroz",
    "Ravenna Maria Barroso Gomes",
    "Rayane Batista Saboia",
    "Rebeca Mesquita",
    "Renan Dos Santos Nogueira",
    "Renata Cristina Façanha De Menezes",
    "Renata De Alencar Pinheiro", "MARCELA FERREIRA DE OLIVEIRA PINTO" ,
    "Renata Pinheiro",
    "Renato Tartuce",
    "Riamburgo Ximenes",
    "Ricardo Borges",
    "Ricardo Castro",
    "Ricardo Montefusco",
    "Ricardo Vasconcelos",
    "Rita Edvirgem Carvalho Fernandes",
    "Roberta Arruda",
    "Roberta da Rocha Amadei",
    "Roberta Jereissate Ary Carneiro",
    "Roberta Lessa",
    "Roberta Sieny",
    "Roberta Teles",
    "Rodolfo Leitão",
    "Rodrigo Barbosa",
    "Rodrigo Martins de Paica Sales",
    "Rodrigo Xerek",
    "Roger Barbosa Mesquita",
    "Rosada Pinheiro",
    "Rosangela Figueiredo",
    "Rosimar Fernandes",
    "Rosseline Santa Rosa",
    "Salusa Rosas",
    "Samuel Dias",
    "Samylla Cardoso Tavares",
    "Sara Ferreira",
    "Sarah Barroso Ribeiro Facó",
    "Sheryda Lorrayna",
    "Shirley Lessa Seba Almeida",
    "Silvana Pinheiro de Oliveira",
    "Sophia da Cunha Geraldo Costeski",
    "Stella Colares",
    "Susana Pompeu Saraiva Ribeiro",
    "Tarsila Gabriele Pereira",
    "Tatila Emanuela Melo e Almeida",
    "Tereza Lucia",
    "Thais Macedo Feijó Lima",
    "Themisa Pimentel",
    "Thiago Sobreira",
    "Thiago Vasconcelos",
    "Thiago Vieira",
    "Thiziane Palacio",
    "Tif TESTE",
    "Vanessa Brito",
    "Vanessa Veras",
    "Vanessa Viana",
    "Veridiana Dias Monteiro",
    "Verônica Bringel de Oliveira Torres",
    "Vicente Anderson Paz Sales",
    "Victor Afonso",
    "Victor Ibiapina",
    "Victor Montenegro",
    "Victor Salmer",
    "Victoria Correia dos Santos",
    "Vidermânia Duarte",
    "ville",
    "Vitória Karoline De Lacerda",
    "Wládia Lélis",
    "Wolney Mattos Oliveira",
    "Yanna Lia Gadelha Moura",
    "Ylmara Ivna",
    "Yvie Ana Alves De Queiroz",
    "Aline Gradvohi",
    "Mariana Oliveira",
    "Taina Barbara",
    "Flávia Soares Unneberg",
    "Milena Vasconcelos Aguiar",
    "Calina Accioly",
    "Carmelo Queiroz",
    "Hany Souza",
    "Karina Bernardes",
    "Ana Catani",
    "Hildo Costa",
    "Samara Ribeiro Moura",
    "Charles Gurgel Martins",
    "Paulo Rolim",
    "Janaina Ramos",
    "Denise Brandão",
    "Heitor Lopes",
    "Leandro Terto",
    "Jorge Pedro",
    "Ryan Oliveira do Nascimento",
    "Lidiane Oliveira",
    "Diego Antunes Silveira",
    "Milena Barros",
    "Rubens dos Santos Gomes",
    "Ana Carolina da Rocha",
    "Larissa Sampaio Monteiro",
    "Rafhael cunha",
    "Natasha Queiroz",
    "Rafaella Melo",
    "Ana Leda Neves",
    "Yuri Aragão Alves",
    "Zack Haward",
    "Issa K",
    "Nat Viana",
    "Geiza Claudia Mota Feitosa",
    "Leandro Vidal",
    "Maria Inês Costa",
    "Ricardo Tavares",
    "João Almeida",
    "Iluska Suassuna",
    "Gabriel Almeida",
    "Erica Tavares",
    "Caio Rocha",
    "Divane Feitosa Mariz",
    "Kersia Silva",
    "Thaina Cavalcante",
    "Victor Gomes",
    "Thamires Ther",
    "Bruno Dias",
    "Cristiane Mota de Holanda Pereira",
    "Moesio Cavalcante",
    "JULIA DAVILA SOUSA",
    "JOSE FELIPE DE ALMEIDA",
    "Steffany Gadelha",
    "RAIZA MELO",
    "LAZARO MENDES MAIA NETO",
    "ANA CAROLINE",
    "MOHAMED BLUE HEART",
    "JULIANA GONÇALVES LEITAO",
    "ANA CAROLINE SOUSA ALMEIDA MAIA",
    "JOHN BRITO",
    "Paloma Sampaio",
    "GABRIELA SOBRAL",
    "LILIAM MAIA DE MORAES SALES",
    "ALBERTO BIAL",
    "WALMIR DE CASTRO JUNIOR",
    "LUCAS RIGONATI",
    "ANA DE FATIMA URANO CARVALHO",
    "LUANE THAYNA DOS REIS CUNHA",
    "MARCELO PEREIRA D'ALENCAR",
    "VICTOR GUSSÃO MEDEIROS",
    "ADRIANA SAMPAIO LIMA",
    "NATALIA DE ARAGÃO PINTO",
    "FERNANDO MULLER",
    "TELURIO FREIRE",
    "REBECA MOURAO",
    "PATRICIA TAMAR SOARES",
    "Danielli Muller",
    "Emanuel Linhares",
    "Tiara Lopes dos Santos",
    "Igor Bastos Pedras",
    "Jessika Garrido Barbosa Marzola",
    "Elizabeth Barbosa de Melo",
    "Rafael Nasser",
    "Aline Fernandes",
    "Katiusia Bezerra viana",
    "Ramon Marques",
    "Adil Dallago",
    "Barbara Gomes",
    "Marlon Lima",
    "Stephane Blanc",
    "Isis Rodrigues Chidic",
    "Mário Nelson",
    "Christine Branco",
    "Maria Eliete de Almeida Cordeiro",
    "Rebeca Uchoa Saraiva",
    "Lorena Araujo Vasconcelos",
    "Iandara martins",
    "Diogo Neves",
    "Morgana Pordeus",
    "Laina Vieira",
    "Pedro Lima",
    "Larissa Leitão", "CLARISSA DEUSDARA",
    "Nicolas Ximenes",
    "Vitória Amelia de Carvalho da Costa",
    "Bruno Lopes",
    "VItória Regia Taveres Viana",
    "Tales Sampaio",
    "Daniel Grangeiro",
    "Laiane Grangeiro",
    "Thassia Feijo Dantas",
    "Francisco Aldair",
    "Maria Cristina FIgueiredo Sobral",
    "Erica Gusmão",
    "Aline Holanda",
    "Luana Braga",
    "Catarina Silva",
    "Marina Brasileiro",
    "Caroline Prover",
    "Samia Lima",
    "Priscili Fortaleza",
    "Mirela Lima",
    "Leticia Lima (Mirela)",
    "Lívia Lessa",
    "Silvana Parente Fernandes",
    "Ana Elizabeth",
    "Iara Randal",
    "Larissa Marques Lima",
    "Etevaldo Nogueira",
    "Bruna Dote",
    "Diana cunha",
    "Neuza Moura",
    "Sarah Gadelha Molta",
    "Michele Montier",
    "Beatriz Rodrigues Andrade",
    "Karine Martinez",
    "Raimundo Pereira Martins",
    "Ana Beatriz Landin",
    "Thalita Moraes",
    "Maria do Socorro Costa Dias",
    "Ana Vitoria Delfino (Estetica)",
    "Virgínia Araújo",
    "Rebecca Freire",
    "stefanie woolv",
    "Lucia Maria Aragão Soares",
    "Raissa Cavalcante Vaconscelos",
    "Juliana Melo",
    "Roberta Amadei",
    "Thereza Rocha",
    "Rachel Guimaraes",
    "Raimundo Dias",
    "Fabio de farias feitosa",
    "Iara Andrade",
    "Vera Lucia Brito",
    "Camile Firme",
    "Yuri Melo Araujo",
    "Lara Andrade Vieira",
    "André Sena",
    "Maria Selma",
    "Arthur Feijó",
    "Giselle Ribeiro",
    "Brunno de medeiros carvalho barreto",
    "Cezar Fernandes Brito",
    "Samia Ximenes",
    "Laurez Veba",
    "Ana carolina dos Santos Silva",
    "Caroline Feitosa",
    "Julia Veba",
    "Neusa Moura",
    "João Paulo Mattos",
    "Sâmia Oliveira Freitas diógenes",
    "Thiago Leitão",
    "Fernanda Delgado",
    "Douglas Camargo",
    "Juliana Bastos",
    "Grace Kelly Lopes",
    "Marcelo Faria",
    "Nayane Mota",
    "Karla Cunha",
    "Luciana Russo Leal",
    "Roberta Ary Carneiro",
    "Francisca Viviane Teixeira",
    "José Edson Cavalcante",
    "Samuel menezes pimenta",
    "Davi Holanda Diorgenes",
    "Adilson Freitas",
    "Sandro Rodrigues",
    "Ana Vladia Feitosa",
    "Beda Cesar Facó",
    "Ruth Maria dos Snatos",
    "Lívia Teles Nascimento",
    "Príscila Oliveira da Silveira",
    "Dana Cortez",
    "Gabriella Queiroga",
    "Dayana Serpa Diniz Baima",
    "Clauber de Moura Baima",
    "Ticiane Lima",
    "Jéssica Marques",
    "Emilia Campelo Borges",
    "Frida Doria",
    "Humberto de Aráujo Rocha",
    "Solange",
    "Ney Conceição",
    "Josy Portela",
    "Iara Marques",
    "Rosalia Sampaio Mendes",
    "Ronielton da Costa de Sousa",
    "Marcus Amaro Marques",
    "Bárbara Vitorino de Silva",
    "Rejane Nogueira Pamplona",
    "Rafael Lino Façanha Soler",
    "Antonio Leandro Silva de Sousa",
    "Ane Patricia",
    "Ari Martins",
    "Ana Luísa Menezes",
    "Maria Isabel Liborio",
    "Raquel Andrade Sales",
    "Beatriz Gurgel",
    "Regina Celia Cardoso",
    "Duan Barros",
    "Diego Fernandes Cavalcante",
    "Caio Noah Teixeira",
    "André Vargas",
    "Lurdiane Rodrigues Almeida",
    "Camila Oliveira Bandini",
    "Karol",
    "Alice de Melo Ribeiro",
    "Rodrigo de Lima Ferreira",
    "Miguel Kikuchi",
    "Leonardo marinho",
    "Elaine Pereira",
    "Eline Rabelo Frota",
    "Mateus da silva Maciel",
    "Cinthia Alencar",
    "Joana Parente",
    "Luciana Lopes de Castro",
    "Maria Clara Moreira",
    "Débora Dias Bandeira",
    "Deusarina Dias Bandeira",
    "Barbara Pasquera",
    "Cleidson Alcântara",
    "Kellyana Miranda",
    "Shalon Miranda",
    "Raquel Albuquerque",
    "Jani da Silva",
    "Luiz Viana",
    "Ísis Gomes",
    "Vanessa Gomes",
    "Natalia Barros",
    "Fernando Cesar",
    "Amanda Holanda Baia",
    "Mariana Karbage",
    "Luiza Amélia",
    "Roberlene Rodrigues",
    "Ruric Chunacero",
    "Sara Farias",
    "Aline Martins",
    "Monica Mendes",
    "Aurea Lucia",
    "Maria Ignes",
    "Renato Maia Nogueira",
    "Bruno Veloso",
    "Michele Ribeiro",
    "Rafael Flusa",
    "Isabelle fragoso",
    "Kianna Ivi",
    "Idaliana teixeira",
    "Lidia aurelia",
    "Malu Wheyne",
    "Armando Sergio",
    "Daniel Soares de Sousa",
    "Gabriela Holanda",
    "Rebeca Alves",
    "Clara de vasconcelos Lopes",
    "Livia Vidal",
    "Anita lima",
    "Estefânia Cavalcante",
    "Clarissa Cabral",
    "Gabrielle Oliveira Kanheski",
    "Catarina Vasconcelos de Queiroz",
    "Rafael Kanheski",
    "Alessandra Coelho",
    "João Pedro Borges",
    "Rebeca Araújo",
    "Renatynha Xavier",
    "Carolina Fonteneles",
    "Stavian Cann",
    "Juliana de Sousa Araújo",
    "Adriano Holanda Viana",
    "Hino Mastroto",
    "Karinna Matos",
    "Caio Burgos",
    "Valeria Cruz",
    "Cristiane Mota de Holanda Pereira",
    "Henrique Safar Borges",
    "Ilka Marques",
    "Jacinta Maria",
    "Flavia de Araújo Barbosa",
    "Diego Santana",
    "Luisa Furtado",
    "Liana",
    "Sabrina Lima",
    "Emylly Cardoso",
    "Idalina Jessica",
    "Walter Mota",
    "Vitoria Costa",
    "Cássia Dummar",
    "Paulo Jose Benevides",
    "Vitoria Amelia de Carvalho",
    "Lia Mara Bernardes",
    "Camila Sousa Schultz",
    "Giselle Bastos",
    "Laecio Mauricio(estetica)",
    "Yasmênia Lima",
    "Julierme Sena(estetica)",
    "Valdemir Rolin de Souza",
    "Gabriel Freire",
    "Leticia Coelho Cavalcante",
    "David Gondim",
    "Mariana Ferrer",
    "Kikuhe Shimabukuro",
    "Thaianne Cassed da Silva",
    "Raquel Craveiro",
    "Pedro Lira",
    "Erika Perdigão",
    "Aline Diogenes",
    "Antonio Jardel Freire",
    "Hugo Vasconcelos",
    "Cleber Vidal",
    "Lima Junior",
    "Debora Barros",
    "Antonio Pierre",
    "Ana Clara Paiva",
    "Cristian Teixeira",
    "Sarah Gracielly Sena Sousa",
    "Suelen Lopes",
    "Edilene Vitorino",
    "Jessica Veraz",
    "Gilmario Texeira",
    "Vanessa Veraz",
    "Andréa da Nobrega",
    "Davi Mota",
    "Bruna Coimbra",
    "Daniel Coimbra",
    "Gessica",
    "Miguel Kikuchi",
    "Luisa Rocha",
    "Liana Cunha",
    "Rafael",
    "Waleska Catonio",
    "João Ribeiro Lima",
    "Clarissa Deus Dara Gomes",
    "Raquel Ximenes",
    "Armando",
    "Amaral Catonio",
    "Rafhael Alves da Costa(estetica)",
    "Lara Correa",
    "André Vinícius",
    "Ruy Frota Barbosa(estetica)",
    "Natália de Aragão Pinto(estetica)",
    "Paula Carls(estetica)",
    "Olivia Mota",
    "Hannah Barbosa",
    "Thiago Roberto Coradi",
    "Perola Holli Gsworth",
    "Roger Etherington",
    "Carlos de Albuquerque Herique",
    "Deborah Hahn",
    "jessyka Lima",
    "Marcio de Queiroz Fernandes",
    "Stephanie de Miranda Melo",
    "Fernando Lima",
    "Cristiano Varzim",
    "Fernando Ximenes",
    "Maria Messer",
    "Renata Queiroz",
    "Lidia Isabelle Oliveira",
    "Pietra feijão",
    "Francisco de Morais",
    "Diego Meneses",
    "Livia Botelho",
    "Morgana Cleria Braga",
    "Jamile Olinda",
    "Otavio da Rosa",
    "Nelson Bruno Valença",
    "Geanne Sales",
    "Vanessa Marques",
    "Flavia Bacelar",
    "Daniele Ferreira dos Santos",
    "Marília Lima",
    "Camila de Carvalho",
    "Ana Carolina Agapito",
    "Johnatan Pereira",
    "Caio Batista",
    "Victor Perfecto",
    "Renan",
    "Camila Arruda",
    "Mariana Leitão",
    "Roberta Costa",
    "Lia Alves",
    "Claudio Galantes",
    "Patricia Oliveira",
    "Natalia cavalcante",
    "Sofia Vitorino",
    "Cristiane Santana",
    "Geovana",
    "Idalina Gurgel",
    "Francelho Magalhaes",
    "Leonidas Rosendo",
    "Amanda Lopes",
    "Bruce Brandão",
    "Marcela Altale",
    "Karla Lays",
    "Claudina Silva",
    "Luana Albuquerque",
    "Ana Maria Araujo",
    "clairy Serrath",
    "Bruno Cardoso",
    "Marilia Dias",
    "Rômulo Cesar",
    "Isabelle Cesar",
    "Kelly Mota",
    "Victor Afonso (2)",
    "Vanessa Veras (2)",
    "Helda Kelly",
    "Lia Lopes",
    "Fabilene Silva",
    "Nayara Mariconi",
    "Carlos Magno",
    "Juliana Montenegro",
    "Isabelle Montenegro",
    "Golvida Lila",
    "Bruna Almeida",
    "Fabio Gaspar",
    "Naiara Menezes",
    "Patrícia Borges",
    "Larissa meireles Parente",
    "Kennedy Cavalcante",
    "Cristiane Albuquerque",
    "Roberto Correira",
    "Gabriela Picanço",
    "Maria Taianay Gonçalo",
    "Adriana Arrigoni",
    "Ana Clecia",
    "Rebeka Aragão",
    "Ivana Soares",
    "julia gonçalves",
    "Alana Rodrigues",
    "Lina Arrais",
    "Pedro Viriato",
    "Fernando Bezerra",
    "Marcia Cristina Leitão Pimentel",
    "Luciana Barroso",
    "Aline Yang",
    "Raquel da Silva Soares",
    "Francisco Dias Martins",
    "Neyara Bessa",
    "Francisca Israelly Viana Bezerra",
    "Natacha Medeiros",
    "juliana sampaio de alencar",
    "Jania Pinho",
    "Raul Nogueira Bessa",
    "Ana Carolina Racomonte Capelo",
    "Livia Sousa",
    "João Paulo Santiago",
    "Mailton Oliveira de Arruda",
    "Beni Feitosa Neto",
    "Daniella da Silva Duarte",
    "Raquel Dantas do Amaral",
    "Jorge Adrian Goes",
    "Jady Lima",
    "Georgia Cunha",
    "Maria Carolina Arrai Pedras",
    "Leonardo PorDeus Barroso",
    "Rebeca Nocentini",
    "cecilia barroso(estetica)",
    "Alexandre Atílio Ramos de Alencar",
    "Fernanda Alves castro",
    "Ursula Saboia",
    "mathieu wrobleski",
    "Nathan Miranda",
    "Ariedson Rocha",
    "Eveline Gentil",
    "Rhavena Sá",
    "Fabia Melo",
    "Irene Isaias",
    "Maria Girão silva",
    "Karla Colasso",
    "Vicente Paulo Vasco junior",
    "Lisiane Cysne(estetica)",
    "Fernanda Freire Rodrigues",
    "Lívia Bessa Gomes",
    "Diego Vasconcelos",
    "Clarice Maciel",
    "Leny Rosa Correia",
    "Arthur Alencar",
    "Rita Pedrosa",
    "Roberto Terra",
    "Juliana Carvalho Mesquita",
    "Steven Jensen",
    "Lisiane Cysne",
    "Josy Meire Sales",
    "Bruna Pereira",
    "Leticia Gomes",
    "Lissa Vale",
    "Ravelly bezerra de meneses marques",
    "Jeannine Maria Limaverde Freitas",
    "Giselle Kawano",
    "Maria de Fátima Bastos Nóbrega",
    "Thais",
    "Ares",
    "Lara Guerra Lucena Matias Alencar",
    "Mônica Fonseca",
    "Stephanie Martins",
    "Leno Dos Santos",
    "Anelize Rosangela",
    "Elaine Jereissati",
    "Leonardo Carls",
    "Renan Delfino",
    "Matheus Da Costa",
    "Sofia Albuquerque",
    "Priscilla Rodrigues",
    "Tatiane Roseno",
    "Viviane Viana",
    "Patricia",
    "Tassia Gabriela",
    "Sorany Alcantara", "Rafaela Keyla Gomes ," ,
    "Sarah Portela",
    "Rafaela Leitão",
    "Marcelle Noronha",
    "Victoria Medeiros",
    "Ingrid Solheiro",
    "Livia Teles",
    "Lucca Gian Picco",
    "Vonia Maria",
    "Rebeca Gregory",
    "Eirles Lemos",
    "Maria da Conceição",
    "Joyce Coling",
    "Leidiane Pinheiro",
    "Georgia Catunda",
    "Giovanna Matias Duarte",
    "João Breno Sampaio",
    "Lisani",
    "João Neto",
    "Blue Birdie Weyne Saisho",
    "Mayara Magalhães Martins",
    "Rebeca Passamani",
    "José Neto",
    "Emanuel Sátiro",
    "Ana Ribeiro",
    "Maira Cunha",
    "João Deveikis Neto",
    "Luis Fernando",
    "Ticiana Gomes Cavalcante",
    "Humberto Nunes",
    "Leticia Peixoto",
    "Lara",
    "Nicolle Ingra",
    "Felipe Franco",
    "Karol Matias",
    "Jessika Karla Sá",
    "Deoclides Melo",
    "Abel Henrique Cavalcante",
    "Beatriz Vasconcelos Silveira",
    "Geraldo Silva Neto",
    "Rebeca",
    "Graziella Brito",
    "Marcio Alcantara",
    "Maina Araújo Sampaio",
    "Lunna Lima",
    "Jorge Salgueiro",
    "John Unneberg",
    "Simião Oliveira",
    "Aline Vilar",
    "Deilane Amorim",
    "Julia Pereira Henrique",
    "Elida Cristina",
    "Lis Feitosa",
    "Shayanne Feitosa Andrade",
    "Vera",
    "Jean Carlos",
    "IGOR BASTO",
    "JOSE WILSON RICARTE JOSINO",
    "Natalia Tavora",
    "Marcio Pires del Picchia",
    "Marcio Marzola",
    "Paloma Sampaio",
    "Silvana Parente Fernandes",
    "Thiago Prada Correia Lima",
    "Adah Emille Guilherme",
    "Rafael Quariguasy",
    "Isadora Quariguasy Veras Leitão",
    "Isabelle Duque",
    "Paulo Ferreira",
    "Victoria Vasconcelos",
    "Francisco Ary Carneiro",
    "Manuela Sales",
    "Ingra Marie Furtado",
    "Mariana Furtado",
    "Cybele Viana",
    "Marianna Rodrigues",
    "Rafaela Feitosa",
    "Raul Corrêa Guimarães",
    "Emilian Bredael",
    "Amanda Karine Silveira Campelo",
    "Graça Façanha",
    "Saul Stefano Rodrigo Martins",
    "Tatiana Sampaio",
    "Romulo Calado",
    "Aline Bentemuller",
    "Claudiana",
    "Julia Almeida",
    "Antonia Juliana",
    "Luiza Albuquerque",
    "Maria Eduarda Galvão",
    "Rayanne Melo",
    "Naiara Andrade",
    "Andre de cavalcante",
    "Bruno Pinheiro maia torquato",
    "Eduardo Cerqueira Cunha mascarenhas",
    "Larissa Gurgel",
    "Mariana Saraiva Macambyra",
    "Juliana Bezerra",
    "Mardonio Mateus",
    "João (Brena)",
    "Luciana Braga",
    "Samuel Nunes Limeira",
    "Natalia Ferreira Lima",
    "Yasmin Moreira",
    "Camila Pessoa",
    "Cida Soares",
    "Ordaldo Moreira",
    "Germana Moreira",
    "Nicodemos Maia",
    "Fernanda Esteves",
    "Raissa Coutinho",
    "Flavia Brito",
    "Joselidia Coelho",
    "Renata Rocha",
    "Ana Beatriz Henrique",
    "Leticia Ferreira",
    "Yasmin Ferreira",
    "Raiana Pereira",
    "manuela palacio",
    "Monique Brito",
    "Fabio yang",
    "clarissa barroso",
    "tony yang",
    "clairy serrath",
    "juliana sampaio",
    "Kaoanne Santos",
    "Fernanda Freitas",
    "Rogerio Arruda Gonçalves",
    "Clayton Henrique Rocha Macedo",
    "Marina Studart",
    "Roberlene rodrigues",
    "Cristian Teixeira",
]

    # Adicione mais clientes conforme necessário


procedimentos_predefinidos = [ "Relaxante 1:30h" ,"Relaxante 1h" ,
"Crânio Facial" ,
"Spa Dos Pés" ,
"Pedras Quentes  1h" ,
"Pedras Quentes 1:30h",
"Vela Quente 1h" ,
"Vela Quente 1:30h" ,
"Relaxante Com Ventosa 1h" ,
"Relaxante Com Ventosa 1:30h" ,
"Revitalizante 1h" ,
"Revitalizante 1:30h" ,
"Desportiva 1h" ,
"Desportiva 1:30H" ,
"Drenagem 1H" ,
"Drenagem 1:30H" ,
"Mandala Slin 1H" ,
"Mandala Slin 1:30H" ,
"Magic Detox" ,
"Magic Detox Face" ,
"Momento Mandala" ,
"Mandala Skin" ,
"Mandala Luxo" ,
"Day Spa" ,
"Avaliação" ,
"Reavaliação" ,
"Power Detox" ,
"Detox Slim" ,
"Limpeza De Pele (Costas)" ,
"Lipocavitação (Áreas)" ,
"Radiofrequencia (Áreas)",
"Peeling Clareador",
"Hidratação / Revitalização" ,
"Limpeza De Pele Simples" ,
"Limpeza De Pele Personalizada" ,
"Micro Agulhamento" ,
"Radiofrequencia" ,
"Depilação  - Abdome Completo" ,
"Depilação Laser - Abdome" ,
"Depilação Laser - Axila" ,
"Depilação Laser - Braço Completo" ,
"Depilação Laser - Contorno Completo" ,
"Depilação Laser - Contorno Simples" ,
"Depilação Laser - Costas" ,
"Depilação Laser - Ombros" ,
"Depilação Laser - Linha Alba" ,
"Depilação Laser - Mãos" ,
"Depilação Laser - Meia Perna (Panturrilha)" ,
"Depilação Laser - Meia Perna (Coxa)" ,
"Depilação Laser - Glúteos" ,
"Depilação Laser - Perianal",
"Depilação Laser - Peito" ,
"Depilação Laser - Pernas",
"Depilação Laser - Pés",
"Depilação Laser - Buço",
"Depilação Laser - Faixa De Barba",
"Depilação Laser - Cabeça",
"Depilação Laser - Rosto Completo" ,
"Depilação Laser - Orelha",
"Depilação Laser - Costeletas",
"Liberação 1H",
"Liberação 1:30H" ,
"Vip 1H" ,
"Combo Paizão" ,
"Relaxante 30Min",
"Combo Super Pai",
"Combo Relax",
"Depilação Laser - Aureola",
"Combo Cristal",
"Combo Diamante",
"Combo Casal Vip",
"1H Modeladora",
"Ultra Cavitação",
"Excepcional",
"Revitalização 2H",
"Banho Relaxante",
"Combo Carinho",
"combo Verão",
"ultra localizada",
"30 min modeladora",
"Drenagem 30MIN",
"Casal radiante",

]



# Variáveis globais
clientes_conhecidos = clientes_predefinidos.copy()
procedimentos_validos = procedimentos_predefinidos.copy()


def atualizar_clientes_novos(caminho_planilha, clientes_conhecidos):
    """
    Verifica a presença de clientes novos (marcados com 'cli') na planilha e os adiciona temporariamente à lista de clientes.
    """
    # Abrir a planilha para leitura
    workbook = load_workbook(caminho_planilha, data_only=True)

    # Iterar por todas as abas e linhas para buscar novos clientes
    novos_clientes = set()
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str) and cell.strip().lower().startswith("cli "):
                    novo_cliente = cell.strip()[4:]  # Remove 'cli ' para obter apenas o nome do cliente
                    novos_clientes.add(novo_cliente)

    workbook.close()
    # Adicionar novos clientes à lista de clientes conhecidos temporariamente
    clientes_conhecidos.extend(novos_clientes)
    clientes_conhecidos = list(set(clientes_conhecidos))  # Remover duplicatas

# Definir a cor vermelha (RGB)
COR_VERMELHA = 'FFFF0000'  # Código RGB para vermelho
COR_AZUL = 'FF0000FF'  # Código RGB para azul  # Código RGB para amarelo
def obter_valor_celula(ws, row, column):
    """
    Obtém o valor de uma célula, considerando células mescladas.
    Se a célula estiver mesclada, retorna o valor da célula mesclada.
    """
    cell = ws.cell(row=row, column=column)
    if cell.value is not None:
        return cell.value
    else:
        # Verificar se a célula está em um intervalo mesclado
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Obter a célula superior esquerda do intervalo mesclado
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                return top_left_cell.value
        return None

def verificar_cor_celula(celula):
    """
    Verifica se a célula está preenchida com a cor vermelha ou amarela.
    Retorna 'Vermelha', 'Amarela' ou None.
    """
    fill = celula.fill
    if fill and hasattr(fill, 'fgColor'):
        cor = fill.fgColor.rgb
        if cor == COR_VERMELHA:
            return 'Vermelha'
        elif cor == COR_AZUL:
            return 'Azul'
    elif fill and hasattr(fill, 'start_color'):
        cor = fill.start_color.index
        if cor == COR_VERMELHA:
            return 'Vermelha'
        elif cor == COR_AZUL:
            return 'Azul'
    return None


def converter_horario(horario_celula):
    """
    Converte o valor da célula de horário em um objeto datetime.
    Retorna None se o horário for inválido ou ausente.
    """
    if horario_celula is None:
        logging.debug("Célula de horário está vazia.")
        return None

    if isinstance(horario_celula, float):
        # Horário representado como número de dias desde 30 de dezembro de 1899
        try:
            return from_excel(horario_celula)
        except Exception as e:
            logging.error(f"Erro ao converter horário do tipo float: {horario_celula}. Erro: {e}")
            return None
    elif isinstance(horario_celula, time):
        return datetime.combine(datetime.today(), horario_celula)
    elif isinstance(horario_celula, datetime):
        return horario_celula
    elif isinstance(horario_celula, str):
        horario_celula_original = horario_celula  # Para log
        # Substituir ';' por ':' e remover espaços
        horario_celula = horario_celula.replace(';', ':').strip()

        # Log após substituição
        logging.debug(f"Convertendo horário: Original='{horario_celula_original}' => Substituído='{horario_celula}'")

        # Remover quaisquer caracteres não numéricos e ':', exceto os necessários para o formato
        horario_celula = re.sub(r'[^\d:]', '', horario_celula)

        # Verificar se o formato está correto após limpeza
        if not re.match(r'^\d{2}:\d{2}(:\d{2})?$', horario_celula):
            logging.error(f"Formato de horário inesperado após substituição e limpeza: '{horario_celula_original}'.")
            return None

        formatos = ["%H:%M", "%H:%M:%S"]
        for formato in formatos:
            try:
                return datetime.strptime(horario_celula, formato)
            except ValueError:
                continue
        logging.error(f"Horário inválido após substituição e verificação de formato: '{horario_celula}'.")
        return None
    else:
        logging.warning(f"Tipo de horário não suportado: {type(horario_celula)}.")
        return None




def extrair_forma_pagamento(texto, threshold=80):
    """
    Extrai a forma de pagamento da célula, encontrando 'deb' para débito, 'cred' para crédito,
    e 'pix' para pagamentos via Pix, independentemente de espaços extras, pontuações e diferenciação de maiúsculas/minúsculas.
    """
    # Adiciona "pix" ao dicionário de formas de pagamento
    formas_pagamento = {
        "deb": "Débito",
        "cred": "Crédito",
        "pix": "Pix"
    }

    # Normalização do texto: Remove caracteres especiais, converte para minúsculas
    texto_normalizado = re.sub(r'[^a-zA-Z0-9\s]', ' ',
                               texto.lower())  # Remove caracteres especiais e converte para minúscula
    palavras_texto = re.split(r'\s+', texto_normalizado)  # Divide o texto em palavras ignorando múltiplos espaços

    for palavra in palavras_texto:
        melhor_ajuste = process.extractOne(palavra, formas_pagamento.keys(), scorer=fuzz.WRatio, score_cutoff=threshold)

        if melhor_ajuste:
            chave = melhor_ajuste[0]
            return formas_pagamento[chave]  # Retorna "Débito", "Crédito" ou "Pix" com base no melhor ajuste encontrado

    return "Forma de pagamento não especificada"


def extrair_primeira_vez(texto):
    """
    Extrai o nome ou indicação de 'primeira vez' no texto.
    Retorna o nome ou valor extraído após a expressão 'primeira vez'.
    """
    # Procurar por 'primeira vez' seguido de algum texto descritivo
    match = re.search(r'\bprimeira vez[:\s]*([A-Za-zÀ-ÖØ-öø-ÿ ]+)', texto, re.IGNORECASE)
    if match:
        return match.group(1).strip()  # Extrai o texto após 'primeira vez'
    return None

def extrair_detalhes_bloco(bloco_texto, funcionario, clientes_conhecidos, procedimentos_validos):
    """
    Extrai os detalhes do bloco de texto, utilizando separadores para segmentar as informações.
    """
    if not bloco_texto:
        logging.debug("Bloco de texto vazio, nenhum detalhe a extrair.")
        return {
            "cliente": None,
            "estetica": "Não",
            "procedimentos": [],
            "telefone": "",
            "valor": None,
            "voucher": "Não",
            "plano": "Não",
            "vem_com": "",
            "obs": "",
            "pln": "Não",
            "observacao": "",
            "pref": "",
            "primeira_vez": "",
            "acrescentar_30_min": False  # Inicialização da nova chave
        }

    texto = bloco_texto.strip()
    logging.debug(f"Processando bloco de texto: {texto}")

    detalhes = {
        "cliente": None,
        "cli": None,
        "estetica": "Não",
        "procedimentos": [],
        "telefone": "",
        "valor": 0.0,  # Inicialização corrigida
        "voucher": "Não",
        "plano": "Não",
        "vem_com": "",
        "obs": "",
        "pln": "Não",
        "observacao": "",
        "pref": "",
        "primeira_vez": "",
        "acrescentar_30_min": False  # Inicialização da nova chave
    }

    # Extrair telefone antes de qualquer outra coisa
    telefone = extrair_telefone(texto)
    if telefone == "Número incorreto":
        detalhes["observacao"] += "Número incorreto detectado; "  # Marcar na coluna de observação
    else:
        detalhes["telefone"] = telefone
    # Extrair 'primeira vez'
    primeira_vez = extrair_primeira_vez(texto)


    if primeira_vez:
        detalhes["primeira_vez"] = primeira_vez
        # Remove o termo 'primeira vez' do texto para evitar duplicidade
        padrao = r'primeira vez[:\s]+{}'.format(re.escape(primeira_vez))
        texto = re.sub(padrao, '', texto, flags=re.IGNORECASE)

    # Verificar se "PREF" está presente no texto
    if re.search(r'\bPREF\b', texto, re.IGNORECASE):
        detalhes["pref"] = funcionario
        logging.debug(f"'PREF' encontrado. Profissional '{funcionario}' será registrado na coluna 'PREF'.")

    # Extrair procedimentos e remover do texto
    procedimentos = extrair_procedimentos(texto, procedimentos_validos)
    if procedimentos:
        detalhes["procedimentos"].extend(procedimentos)
        for procedimento in procedimentos:
            texto = re.sub(r'\b' + re.escape(procedimento) + r'\b', '', texto, flags=re.IGNORECASE)

    # Definir separadores para segmentar o texto
    separadores = r'[;,|\\/()\[\]\{\}\n\r\t]+|\s{2,}'

    # Dividir o texto com base nos separadores
    partes = re.split(separadores, texto)

    # Remover procedimentos das partes
    partes_sem_procedimentos = [
        parte.strip() for parte in partes if
        parte.strip() and parte.lower() not in [p.lower() for p in procedimentos_validos]
    ]

    # Extrair cliente usando extrair_cli diretamente
    nome_cliente, coluna = extrair_cliente(partes_sem_procedimentos, clientes_conhecidos)
    if coluna == "cli":
        detalhes["cli"] = nome_cliente  # Cliente novo, armazena em 'cli'
    elif coluna == "cliente":
        detalhes["cliente"] = nome_cliente  # Cliente conhecido, armazena em 'cliente'
    # Extrair a presença da palavra 'estetica' no texto
    detalhes["estetica"] = tratar_estetica(texto)

    # Continuar processando as partes restantes
    for parte in partes_sem_procedimentos:
        # Remover "PREF" e "PLN" da parte atual
        parte = re.sub(r'\b(PREF|PLN)\b', '', parte, flags=re.IGNORECASE).strip()

        # Extrair AVL e PLN
        obs, pln = extrair_obs_pln(parte)
        if obs:
            detalhes["obs"] = obs
        if pln == "Sim":
            detalhes["pln"] = pln

        # Verificar voucher
        if "voucher" in parte.lower():
            detalhes["voucher"] = "Sim"


        # Verificar plano (case-insensitive)
        if re.search(r'\bplano\b', parte, re.IGNORECASE):
            detalhes["plano"] = "Sim"

        # Extrair "vem com"
        vem_com = extrair_vem_com(parte)
        if vem_com:
            detalhes["vem_com"] = vem_com

        # Extração do valor monetário
        valor = extrair_valor(parte)
        if valor is not None:
            try:
                detalhes["valor"] += float(valor)  # Incrementar valores
            except ValueError:
                logging.warning(f"Erro ao converter valor: {valor}. Definindo como 0.0.")
                detalhes["valor"] = 0.0

        # Validação final
        detalhes["valor"] = detalhes["valor"] or 0.0  # Garantir que nunca seja None
        # Verificar se há observações
        if "observação" in parte.lower():
            detalhes["observacao"] = parte.replace("observação", "").strip()
            # Verificar se há "+30 min"
            if acrescentar_30_min(parte):
                detalhes["acrescentar_30_min"] = True

    return detalhes


def extrair_telefone(texto):
    """
    Extrai e valida o número de telefone do texto.
    Retorna o número de telefone se válido, ou uma mensagem indicando erro de digitação.
    """
    # Remover espaços extras e normalizar texto
    texto = re.sub(r'\s+', ' ', texto).strip()

    # Padrões para número de telefone brasileiro
    padrao_telefone_br = r'\(?\d{2}\)?\s?\d{4,5}-?\d{4}'

    # Verificar números no padrão brasileiro
    match = re.search(padrao_telefone_br, texto)
    if match:
        telefone = match.group(0)
        # Validar o formato do número brasileiro
        if len(re.sub(r'\D', '', telefone)) in [10, 11]:  # 10 ou 11 dígitos para Brasil
            return telefone
        else:
            return "Número incorreto"  # Possível erro de digitação

    # Verificação adicional para números internacionais
    padrao_telefone_estrangeiro = r'\+\d{1,3}\s?\d{1,14}'  # Padrão básico para internacional
    match_internacional = re.search(padrao_telefone_estrangeiro, texto)
    if match_internacional:
        return match_internacional.group(0)  # Retorna o número internacional encontrado

    # Caso não corresponda a nenhum padrão
    return "Número incorreto"


def extrair_procedimentos(texto, procedimentos_validos, threshold=60):
    """
    Extrai procedimentos do texto usando correspondência aproximada com RapidFuzz.
    Dá prioridade a combos e retorna apenas a melhor correspondência.

    Parâmetros:
    - texto: String contendo o texto a ser analisado.
    - procedimentos_validos: Lista de procedimentos válidos.
    - threshold: Pontuação mínima para considerar uma correspondência (padrão: 80).

    Retorna:
    - Lista com o melhor procedimento encontrado, priorizando combos.
    """
    if not texto:
        return []

    valor_texto = texto.strip()
    melhor_procedimento = None

    # Dicionário para mapear abreviações com tempos específicos
    mapa_abreviacoes = {
        "rlx": ["Relaxante 30min", "Relaxante 1h", "Relaxante 1:30h"],
        "rlx 30 min": ["Relaxante 30min"],
        "rlx 1h": [ "Relaxante 1h",],
        "PEDRAS": ["Pedras Quentes 30min", "Pedras Quentes 1h", "Pedras Quentes 1:30h"],
        "velas": ["Vela Quente 30min", "Vela Quente 1h", "Vela Quente 1:30h"],
        "MODELADORA": ["1H Modeladora", "30 min modeladora"],
        "SLIN": ["Mandala Slin 1H", "Mandala Slin 1:30H"],
        "DETOX": ["Magic Detox", "Magic Detox Face", "Detox Slim"],
        "C\\PEDRAS": ["Pedras Quentes 30min", "Pedras Quentes 1h", "Pedras Quentes 1:30h"]
        # Adicione mais abreviações conforme necessário
    }

    # Definir os separadores para dividir os procedimentos na célula
    separadores = r'[;,|\\/()\[\]\{\}\n\r\t]+|\s{2,}'

    # Dividir o texto com base nos separadores para capturar múltiplos procedimentos
    partes_texto = re.split(separadores, valor_texto)

    for parte in partes_texto:
        parte = parte.strip()
        if not parte:
            continue

        # Verificar se a parte corresponde a uma abreviação
        for abreviacao, variantes in mapa_abreviacoes.items():
            if abreviacao in parte.lower():
                melhor_ajuste = process.extractOne(parte, variantes, scorer=fuzz.WRatio, score_cutoff=threshold)
                if melhor_ajuste:
                    # Atualizar a melhor correspondência encontrada
                    if melhor_procedimento is None or "Combo" in melhor_ajuste[0]:
                        melhor_procedimento = melhor_ajuste[0]
                break

        # Caso não seja uma abreviação conhecida, verificar nos procedimentos válidos
        if melhor_procedimento is None:
            melhor_ajuste = process.extractOne(parte, procedimentos_validos, scorer=fuzz.WRatio, score_cutoff=threshold)
            if melhor_ajuste:
                # Atualizar a melhor correspondência encontrada
                if melhor_procedimento is None or "Combo" in melhor_ajuste[0]:
                    melhor_procedimento = melhor_ajuste[0]

    # Retornar a melhor correspondência encontrada
    return [melhor_procedimento] if melhor_procedimento else []


def extrair_cliente(partes, clientes_conhecidos):
    """
    Extrai o nome do cliente a partir das partes fornecidas.
    - Se o cliente tiver o prefixo "cli", considera-o como um novo cliente e direciona para a coluna 'cli'.
    - Se o cliente já estiver cadastrado em 'clientes_conhecidos', armazena na coluna 'cliente'.
    """
    for parte in partes:
        parte_limpa = parte.strip()

        # Verifica se o cliente é novo usando o prefixo 'cli'
        if parte_limpa.lower().startswith("cli "):
            nome_cliente = parte_limpa[4:].strip()  # Remove o prefixo 'cli '
            return nome_cliente, "cli"  # Indica que é para a coluna 'cli'

        # Verifica se o cliente está na lista de clientes conhecidos
        if parte_limpa in clientes_conhecidos:
            return parte_limpa, "cliente"  # Indica que é para a coluna 'cliente'

    # Caso nenhum cliente seja encontrado
    return None, None


def extrair_cliente(partes, clientes_conhecidos):
    """
    Extrai o nome do cliente a partir das partes fornecidas.
    - Se o cliente tiver o prefixo "cli", considera-o como um novo cliente.
    - Se o cliente já estiver cadastrado em 'clientes_conhecidos', armazena na coluna 'cliente'.
    """
    for parte in partes:
        parte_limpa = parte.strip().lower()

        # Caso o cliente seja novo, identificado pelo prefixo 'cli'
        if parte_limpa.startswith("cli "):
            nome_cliente = parte_limpa[4:].strip()  # Remove o prefixo 'cli '
            return nome_cliente, "cli"  # Retorna o nome e indica que é para a coluna 'cli'

        # Caso o cliente esteja na lista de clientes conhecidos
        for cliente in clientes_conhecidos:
            if cliente.lower() == parte_limpa:
                return cliente, "cliente"  # Retorna o nome e indica a coluna 'cliente'

    # Se não encontrar nenhum cliente válido
    return None, None


def tratar_estetica(texto):
    """
    Verifica se a palavra 'estetica' está presente no texto.
    Retorna 'Sim' se encontrada, caso contrário, retorna 'Não'.
    """
    if "estetica" in texto.lower():
        return "Sim"
    return "Não"


def extrair_obs_pln(texto):
    obs_match = re.search(r'OBS\s*(\d+)', texto, re.IGNORECASE)
    pln_match = re.search(r'PLN', texto, re.IGNORECASE)

    obs = obs_match.group(1).strip() if obs_match else ""
    pln = "Sim" if pln_match else "Não"

    return obs, pln

def extrair_vem_com(texto):
    match = re.search(r'vem com\s+([A-Za-zÀ-ÖØ-öø-ÿ ]+)', texto, re.IGNORECASE)
    return match.group(1).strip() if match else ""

def extrair_valor(texto):
    """
    Extrai e soma valores monetários precedidos por 'R$' no texto.
    Garante que valores sejam identificados mesmo com espaços antes ou depois.
    """
    # Remover barras invertidas e normalizar espaços
    texto = re.sub(r'[\\]', ' ', texto).strip()
    texto = re.sub(r'\s+', ' ', texto)  # Substituir múltiplos espaços por um único

    # Remover números que podem ser telefones
    texto_sem_telefone = re.sub(r"\b\d{2} ?\d{4,5}-?\d{4}\b", "", texto)

    # Padrão para encontrar valores monetários precedidos por 'R$'
    padrao_valor = r'R\$\s*(\d{1,4}(?:[.,]\d{1,2})?)'
    valores = re.findall(padrao_valor, texto_sem_telefone, re.IGNORECASE)

    soma_valores = 0.0

    for valor_str in valores:
        # Substituir vírgulas por ponto para converter em float
        valor_str = valor_str.replace(',', '.')
        try:
            soma_valores += float(valor_str)
        except ValueError:
            continue  # Ignorar valores inválidos

    # Retornar o total como float ou 0.0 se nenhum valor encontrado
    return soma_valores if soma_valores > 0 else 0.0


def acrescentar_30_min(texto):
    """
    Verifica se a string 'bloco +30 min' está presente no texto de forma case-insensitive.
    Se estiver presente, retorna True, indicando que 30 minutos devem ser adicionados ao horário final.
    Caso contrário, retorna False.
    """
    if re.search(r'\bbloco\s*\+30\s*min\b', texto, re.IGNORECASE):
        return True
    return False


def processar_bloco(bloco_celulas, funcionario, ws_novo, sheet_name, horario_inicio_bloco, clientes_conhecidos,
                    procedimentos, cor_bloco):
    """
    Processa um bloco de células de uma determinada cor e combina todos os procedimentos em uma única célula.
    """

    # 1. Concatenar os valores das células do bloco
    valor_bloco = ' '.join(celula['valor'] for celula in bloco_celulas)
    logging.debug(f"Valor concatenado do bloco ({cor_bloco}): {valor_bloco}")

    # 2. Calcular a duração total do bloco (30 minutos por célula)
    duracao_bloco = len(bloco_celulas) * 30  # 30 minutos por célula

    # 3. Calcular o horário de fim do bloco
    horario_fim_bloco = horario_inicio_bloco + timedelta(minutes=duracao_bloco)
    horario_inicio_str = horario_inicio_bloco.strftime("%H:%M")
    horario_fim_str = horario_fim_bloco.strftime("%H:%M")

    logging.debug(f"Horário de início: {horario_inicio_str}, Horário de fim: {horario_fim_str}")


    # 4. Calcular o tempo total em horas e minutos
    tempo_total = horario_fim_bloco - horario_inicio_bloco
    tempo_total = (horario_fim_bloco - horario_inicio_bloco).total_seconds() // 60  # Total em minutos


    # 5. Extrair detalhes do bloco de texto concatenado
    detalhes = extrair_detalhes_bloco(valor_bloco, funcionario, clientes_conhecidos, procedimentos)

    # 6. Extrair a forma de pagamento
    forma_pagamento = extrair_forma_pagamento(valor_bloco)

    # 7. Verificar se deve acrescentar 30 minutos adicionais
    if detalhes["acrescentar_30_min"]:
        horario_fim_bloco += timedelta(minutes=30)
        horario_fim_str = horario_fim_bloco.strftime("%H:%M")  # Atualizar a string de horário fim
        logging.debug(f"Acrescentando 30 minutos ao horário final devido à presença de '+30 min'.")

    # 8. Concatenar procedimentos em uma única string
    procedimentos_unicos = ", ".join(set(detalhes["procedimentos"])) if detalhes["procedimentos"] else ""

    # 9. Preparar o registro para salvar na nova planilha
    registro = [
        str(sheet_name),  # Aba
        str(funcionario),  # Funcionário
        horario_inicio_str,  # Horário Início (já string)
        horario_fim_str,  # Horário Fim (já string)
        detalhes.get("cliente", ""),  # Cliente (mantém vazio se None)
        detalhes.get("estetica", "Não"),  # Estética (padrão "Não")
        ", ".join(detalhes.get("procedimentos", [])),  # Procedimentos combinados em uma única célula
        detalhes.get("telefone", ""),  # Telefone (mantém vazio se None)
        detalhes.get("valor", 0.0),  # Valor (mantém float para exportação correta)
        detalhes.get("voucher", "Não"),  # Voucher (padrão "Não")
        detalhes.get("plano", "Não"),  # Plano (padrão "Não")
        detalhes.get("vem_com", ""),  # Vem Com (mantém vazio se None)
        detalhes.get("obs", ""),  # OBS (mantém vazio se None)
        detalhes.get("pln", "Não"),  # PLN (padrão "Não")
        detalhes.get("observacao", ""),  # Observação (mantém vazio se None)
        detalhes.get("pref", ""),  # PREF (mantém vazio se None)
        detalhes.get("primeira_vez", ""),  # Primeira Vez (mantém vazio se None)
        cor_bloco,  # Cor do Bloco
        detalhes.get("cli", ""),  # Cli (mantém vazio se None)
        detalhes.get("unidade", ""),  # Unidade (mantém vazio se None)
        tempo_total,  # Tempo Total em minutos (mantém numérico)
        forma_pagamento  # Forma de Pagamento
    ]


    # 8. Adicionar o registro à nova planilha
    ws_novo.append(registro)
    logging.info(f"Registro adicionado: {registro}")


def processar_planilha(caminho_planilha, clientes_lista, procedimentos):
    """
    Processa a planilha Excel, identificando blocos de células vermelhas ou amarelas em colunas independentes,
    usando a coluna A como referência para os horários e processando todas as abas.

    Parâmetros:
    - caminho_planilha (str): Caminho para o arquivo Excel a ser processado.
    - clientes_lista (list): Lista de clientes conhecidos para referência.
    - procedimentos (list): Lista de procedimentos válidos para extração.
    """

    # 1. Verificar se o arquivo existe
    if not os.path.exists(caminho_planilha):
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_planilha}")

    # 2. Carregar a planilha original e criar uma nova planilha para salvar os resultados
    wb_original = load_workbook(caminho_planilha, data_only=True)
    wb_novo = Workbook()
    ws_novo = wb_novo.active

    # 3. Adicionar os cabeçalhos na nova planilha
    ws_novo.append([
        "Aba", "Funcionário", "Horário Início", "Horário Fim", "Cliente", "Estética", "Procedimento",
        "Telefone", "Valor", "Voucher", "Plano", "Vem Com", "OBS", "PLN", "Observação", "PREF",
        "Primeira Vez", "Cor" ,"cli"  # Nova Coluna Adicionada
    ])

    # 4. Iterar por cada aba da planilha original
    for sheet_name in wb_original.sheetnames:
        ws_original = wb_original[sheet_name]
        logging.info(f"Processando aba: {sheet_name}")

        # 4.1. Inicializar a variável profissionais
        profissionais = {}

        # 4.2. Utilizar iter_cols para percorrer as colunas, pegando as células das linhas 1 e 2
        for col_cells in ws_original.iter_cols(min_col=2, min_row=1, max_row=2, max_col=ws_original.max_column):
            col_idx = col_cells[0].column  # Obter o índice da coluna
            nomes_encontrados = []

            for cell in col_cells:
                valor = obter_valor_celula(ws_original, cell.row, cell.column)
                if valor:
                    nomes_encontrados.append(str(valor).strip())

            if nomes_encontrados:
                nome_profissional = ' '.join(nomes_encontrados)
                profissionais[col_idx] = nome_profissional

        # 4.3. Se não houver profissionais, adicionar um nome padrão
        if not profissionais:
            profissionais[2] = "Funcionário Desconhecido"
            logging.warning(f"Nenhum profissional encontrado na aba '{sheet_name}'. Usando 'Funcionário Desconhecido'.")

        # 4.4. Para cada profissional, processar as células em sua coluna
        for col, funcionario in profissionais.items():
            linha = 3  # Iniciar na linha 3
            max_linha = ws_original.max_row

            bloco_celulas_vermelhas = []
            bloco_celulas_amarelas = []
            horario_inicio_bloco = None
            cor_atual = None  # Pode ser 'Vermelha', 'Amarela' ou None

            while linha <= max_linha:
                celula = ws_original.cell(row=linha, column=col)
                horario_celula = ws_original.cell(row=linha, column=1).value  # Horário na coluna A

                # Tentar converter o horário, se possível
                horario_inicio = converter_horario(horario_celula)
                if horario_inicio is None:
                    # Se não houver horário, usar um horário padrão ou continuar
                    horario_inicio = datetime.now()
                    logging.warning(f"Horário inválido na aba '{sheet_name}', linha {linha}. Usando horário padrão.")

                # Verificar a cor da célula (Vermelha, Amarela ou outra)
                cor = verificar_cor_celula(celula)

                if cor in ['Vermelha', 'Azul']:
                    if cor_atual != cor:
                        # Se a cor atual mudou, processar o bloco anterior
                        if bloco_celulas_vermelhas:
                            processar_bloco(
                                bloco_celulas_vermelhas, funcionario, ws_novo, sheet_name, horario_inicio_bloco,
                                clientes_lista, procedimentos, 'Vermelha'
                            )
                            bloco_celulas_vermelhas = []
                        if bloco_celulas_amarelas:
                            processar_bloco(
                                bloco_celulas_amarelas, funcionario, ws_novo, sheet_name, horario_inicio_bloco,
                                clientes_lista, procedimentos, 'Azul'
                            )
                            bloco_celulas_amarelas = []
                        # Atualizar a cor atual e iniciar um novo bloco
                        cor_atual = cor
                        bloco_celulas_vermelhas = []
                        bloco_celulas_amarelas = []
                        horario_inicio_bloco = horario_inicio

                    # Adicionar a célula ao bloco correspondente
                    if cor == 'Vermelha':
                        bloco_celulas_vermelhas.append({
                            'linha': linha,
                            'valor': celula.value if celula.value else "",
                            'horario': horario_inicio
                        })
                    elif cor == 'Azul':
                        bloco_celulas_amarelas.append({
                            'linha': linha,
                            'valor': celula.value if celula.value else "",
                            'horario': horario_inicio
                        })
                else:
                    # Se a célula não é vermelha nem amarela, finalizar o bloco atual
                    if cor_atual == 'Vermelha' and bloco_celulas_vermelhas:
                        processar_bloco(
                            bloco_celulas_vermelhas, funcionario, ws_novo, sheet_name, horario_inicio_bloco,
                            clientes_lista, procedimentos, 'Vermelha'
                        )
                        bloco_celulas_vermelhas = []
                    elif cor_atual == 'Azul' and bloco_celulas_amarelas:
                        processar_bloco(
                            bloco_celulas_amarelas, funcionario, ws_novo, sheet_name, horario_inicio_bloco,
                            clientes_lista, procedimentos, 'Azul'
                        )
                        bloco_celulas_amarelas = []
                    cor_atual = None
                    horario_inicio_bloco = None

                linha += 1

            # 4.5. Após terminar as linhas, verificar se ainda está em um bloco
            if cor_atual == 'Vermelha' and bloco_celulas_vermelhas:
                processar_bloco(
                    bloco_celulas_vermelhas, funcionario, ws_novo, sheet_name, horario_inicio_bloco,
                    clientes_lista, procedimentos, 'Vermelha'
                )
            elif cor_atual == 'Azul' and bloco_celulas_amarelas:
                processar_bloco(
                    bloco_celulas_amarelas, funcionario, ws_novo, sheet_name, horario_inicio_bloco,
                    clientes_lista, procedimentos, 'Azul'
                )

    # 5. Salvar a nova planilha com os resultados
    caminho_resultado = caminho_planilha.replace('.xlsx', '-resultado.xlsx')
    wb_novo.save(caminho_resultado)
    logging.info(f"Processamento concluído. Resultados salvos em '{caminho_resultado}'.")


# Configuração básica do logger
logging.basicConfig(level=logging.INFO)

def carregar_clientes():
    """
    Carrega a lista de clientes conhecidos a partir de um arquivo 'clientes.txt'.
    Inicialmente, começa com uma lista pré-definida de clientes.
    Se o arquivo 'clientes.txt' existir, adiciona os clientes contidos nele à lista.
    Remove duplicatas e ordena a lista resultante.
    """
    global clientes_conhecidos  # Indica que estamos modificando a variável global

    # Iniciar com os clientes pré-definidos
    clientes = clientes_predefinidos.copy()

    # Verificar se o arquivo 'clientes.txt' existe
    if os.path.exists('clientes.txt'):
        try:
            with open('clientes.txt', 'r', encoding='utf-8') as f:
                # Ler cada linha, remover espaços em branco e ignorar linhas vazias
                clientes_do_arquivo = [linha.strip() for linha in f if linha.strip()]
                clientes.extend(clientes_do_arquivo)  # Adicionar os clientes do arquivo
            logging.info(f"{len(clientes_do_arquivo)} clientes carregados do arquivo 'clientes.txt'.")
        except Exception as e:
            logging.error(f"Erro ao ler o arquivo 'clientes.txt': {e}")
            messagebox.showerror("Erro", f"Ocorreu um erro ao ler o arquivo 'clientes.txt':\n{e}")

    # Remover duplicatas mantendo a ordem e ordenar alfabeticamente (case-insensitive)
    clientes_conhecidos = sorted(
        set(clientes),
        key=lambda s: s.lower() if isinstance(s, str) else str(s).lower()
    )

    logging.info(f"Total de clientes conhecidos carregados: {len(clientes_conhecidos)}")


def carregar_procedimentos():
    global procedimentos_validos
    procedimentos = procedimentos_predefinidos.copy()  # Inicia com os procedimentos pré-definidos
    if os.path.exists('procedimentos.txt'):
        with open('procedimentos.txt', 'r', encoding='utf-8') as f:
            procedimentos_do_arquivo = [linha.strip() for linha in f if linha.strip()]
            procedimentos.extend(procedimentos_do_arquivo)
    # Remover duplicatas e ordenar
    procedimentos_validos = sorted(set(procedimentos), key=lambda s: s.lower() if isinstance(s, str) else str(s).lower())
    logging.info(f"Procedimentos válidos carregados: {procedimentos_validos}")

def carregar_lista_clientes_txt():
    """
    Abre um diálogo para o usuário selecionar um arquivo TXT contendo a lista de clientes.
    """
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo de clientes",
        filetypes=[("Arquivos de Texto", "*.txt")]
    )
    if caminho_arquivo:
        try:
            with open(caminho_arquivo, 'r', encoding='utf-8') as f:
                clientes_do_arquivo = [linha.strip() for linha in f if linha.strip()]
                # Atualizar a lista de clientes conhecidos
                global clientes_conhecidos
                clientes_conhecidos.extend(clientes_do_arquivo)
                # Remover duplicatas e ordenar
                clientes_conhecidos = sorted(set(clientes_conhecidos), key=lambda s: s.lower() if isinstance(s, str) else str(s).lower())
                messagebox.showinfo("Sucesso", f"{len(clientes_do_arquivo)} clientes carregados com sucesso.")
                logging.info(f"Clientes carregados via arquivo: {clientes_do_arquivo}")
        except Exception as e:
            logging.error(f"Erro ao carregar o arquivo de clientes: {e}")
            messagebox.showerror("Erro", f"Ocorreu um erro ao carregar o arquivo:\n{e}")

def carregar_lista_procedimentos_txt():
    """
    Abre um diálogo para o usuário selecionar um arquivo TXT contendo a lista de procedimentos.
    """
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo de procedimentos",
        filetypes=[("Arquivos de Texto", "*.txt")]
    )
    if caminho_arquivo:
        try:
            with open(caminho_arquivo, 'r', encoding='utf-8') as f:
                procedimentos_do_arquivo = [linha.strip() for linha in f if linha.strip()]
                # Atualizar a lista de procedimentos válidos
                global procedimentos_validos
                procedimentos_validos.extend(procedimentos_do_arquivo)
                # Remover duplicatas e ordenar
                procedimentos_validos = sorted(set(procedimentos_validos), key=lambda s: s.lower() if isinstance(s, str) else str(s).lower())
                messagebox.showinfo("Sucesso", f"{len(procedimentos_do_arquivo)} procedimentos carregados com sucesso.")
                logging.info(f"Procedimentos carregados via arquivo: {procedimentos_do_arquivo}")
        except Exception as e:
            logging.error(f"Erro ao carregar o arquivo de procedimentos: {e}")
            messagebox.showerror("Erro", f"Ocorreu um erro ao carregar o arquivo:\n{e}")

def gerenciar_clientes():
    def carregar_itens():
        return clientes_conhecidos

    def salvar_itens(itens):
        global clientes_conhecidos
        clientes_conhecidos = itens
        # Opcional: salvar em um arquivo, se necessário

    gerenciar_lista(
        titulo_janela="Gerenciar Clientes",
        nome_item_singular="cliente",
        carregar_itens_personalizado=carregar_itens,
        salvar_itens_personalizado=salvar_itens
    )

def gerenciar_procedimentos():
    def carregar_itens():
        return procedimentos_validos

    def salvar_itens(itens):
        global procedimentos_validos
        procedimentos_validos = itens
        # Opcional: salvar em um arquivo, se necessário

    gerenciar_lista(
        titulo_janela="Gerenciar Procedimentos",
        nome_item_singular="procedimento",
        carregar_itens_personalizado=carregar_itens,
        salvar_itens_personalizado=salvar_itens
    )


def gerenciar_lista(titulo_janela, nome_item_singular,
                    carregar_itens_personalizado=None, salvar_itens_personalizado=None):
    """
    Função genérica para gerenciar listas de clientes ou procedimentos.
    """
    janela = tk.Toplevel()
    janela.title(titulo_janela)
    janela.geometry("400x400")

    frame_lista = tk.Frame(janela)
    frame_lista.pack(fill=tk.BOTH, expand=True)

    scrollbar = tk.Scrollbar(frame_lista)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    listbox = tk.Listbox(frame_lista, selectmode=tk.SINGLE)
    listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    listbox.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=listbox.yview)

    def atualizar_lista():
        listbox.delete(0, tk.END)
        itens = carregar_itens_personalizado()
        for item in sorted(itens, key=lambda s: s.lower() if isinstance(s, str) else str(s).lower()):
            listbox.insert(tk.END, item)

    def adicionar_item():
        novo_item = simpledialog.askstring("Adicionar", f"Digite o novo {nome_item_singular}:")
        if novo_item:
            itens = carregar_itens_personalizado()
            itens.append(novo_item.strip())
            salvar_itens_personalizado(itens)
            atualizar_lista()

    def editar_item():
        selecao = listbox.curselection()
        if selecao:
            indice = selecao[0]
            valor_atual = listbox.get(indice)
            novo_valor = simpledialog.askstring("Editar", f"Edite o {nome_item_singular}:", initialvalue=valor_atual)
            if novo_valor:
                itens = carregar_itens_personalizado()
                itens[indice] = novo_valor.strip()
                salvar_itens_personalizado(itens)
                atualizar_lista()

    def remover_item():
        selecao = listbox.curselection()
        if selecao:
            indice = selecao[0]
            itens = carregar_itens_personalizado()
            item_removido = itens.pop(indice)
            salvar_itens_personalizado(itens)
            atualizar_lista()
            messagebox.showinfo("Removido", f"{nome_item_singular.capitalize()} '{item_removido}' removido.")

    botao_adicionar = tk.Button(janela, text="Adicionar", command=adicionar_item)
    botao_adicionar.pack(side=tk.LEFT, padx=5, pady=5)

    botao_editar = tk.Button(janela, text="Editar", command=editar_item)
    botao_editar.pack(side=tk.LEFT, padx=5, pady=5)

    botao_remover = tk.Button(janela, text="Remover", command=remover_item)
    botao_remover.pack(side=tk.LEFT, padx=5, pady=5)

    atualizar_lista()


def selecionar_e_processar_planilha():
    """
    Função para selecionar a planilha, atualizar clientes novos e processá-la.
    """
    caminho_planilha = filedialog.askopenfilename(
        title="Selecione a planilha Excel",
        filetypes=[("Arquivos Excel", "*.xlsx *.xlsm *.xltx *.xltm")]
    )

    if caminho_planilha and os.path.exists(caminho_planilha):
        try:
            # Verificar se as listas de clientes e procedimentos estão carregadas
            if not clientes_conhecidos:
                messagebox.showwarning("Aviso", "A lista de clientes está vazia. Adicione clientes antes de continuar.")
                return
            if not procedimentos_validos:
                messagebox.showwarning("Aviso",
                                       "A lista de procedimentos está vazia. Adicione procedimentos antes de continuar.")
                return

            # Atualizar lista com clientes novos detectados com 'cli'
            atualizar_clientes_novos(caminho_planilha, clientes_conhecidos)

            # Processar a planilha
            processar_planilha(caminho_planilha, clientes_conhecidos, procedimentos_validos)

            messagebox.showinfo("Sucesso",
                                f"Processamento concluído.\nResultados salvos em '{caminho_planilha.replace('.xlsx', '-resultado.xlsx')}'.")
        except Exception as e:
            logging.error(f"Erro ao processar a planilha: {e}")
            messagebox.showerror("Erro", f"Ocorreu um erro ao processar a planilha:\n{e}")
    else:
        messagebox.showwarning("Aviso", "O arquivo selecionado não existe.")
# Criação da janela principal
root = tk.Tk()
root.title("Processador de Planilhas")
root.geometry("400x400")

label_instrucoes = tk.Label(root, text="Selecione uma planilha Excel para processar:", font=("Arial", 12))
label_instrucoes.pack(pady=10)

botao_selecionar = tk.Button(root, text="Selecionar Planilha", command=selecionar_e_processar_planilha, width=25, height=2)
botao_selecionar.pack(pady=5)

botao_carregar_clientes = tk.Button(root, text="Carregar Lista de Clientes", command=carregar_lista_clientes_txt, width=25)
botao_carregar_clientes.pack(pady=5)

botao_gerenciar_clientes = tk.Button(root, text="Gerenciar Clientes", command=gerenciar_clientes, width=25)
botao_gerenciar_clientes.pack(pady=5)

botao_carregar_procedimentos = tk.Button(root, text="Carregar Lista de Procedimentos", command=carregar_lista_procedimentos_txt, width=25)
botao_carregar_procedimentos.pack(pady=5)

botao_gerenciar_procedimentos = tk.Button(root, text="Gerenciar Procedimentos", command=gerenciar_procedimentos, width=25)
botao_gerenciar_procedimentos.pack(pady=5)

# Inicia o loop principal da interface gráfica
root.mainloop()
